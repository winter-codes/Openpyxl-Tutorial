import openpyxl
file_name = 'ttc-streetcar-delay-data-jan-2021-to-august-2021.xlsx'
#Load the workbook
wb = openpyxl.load_workbook(file_name)

#Read sheetnames
#read all sheet names
wb.sheetnames
#read last opened sheet
wb.active
#read a particular sheet by name
wb['Jan 21']
ws = wb['Jan 21']
ws.title
#change sheet name
ws.title = 'January 2021'
wb.save(file_name)

#Read cells
ws = wb['Aug 21']

#Call a single cell
ws['A1']
ws['A1'].value
ws.cell(row=1, column=1)
ws.cell(row=1, column=1).value
ws.cell(row=ws.max_row, column=7)
ws.cell(row=ws.max_row, column=7).value
ws.cell(row=1, column=ws.max_column)
ws.cell(row=1, column=ws.max_column).value

#Call many cells
ws['A1:J1']
ws['A1':'J1']
#entire worksheet
ws['A':'J']
#entire column A
ws['A']
#entire row 2
ws[2]

#Read values from many cells
#read column B
for c in ws['B']:
   print(c.value)

#read row 2
for r in ws[2]:
   print(r.value)

#read entire worksheet values
for i in range(1, ws.max_row):
    row = [cell.value for cell in ws[i]]
    print(row)